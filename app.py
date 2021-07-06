import openpyxl as xl
workbook = xl.load_workbook('transactions.xlsx')
sheet = workbook['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)

for row in range(2, sheet.max_row + 1):
     cell = sheet.cell(row, 3)
     correct_price = cell.value * 0.9
     corrected_price_cell = sheet.cell(row, 4)
     corrected_price_cell.value = correct_price

workbook.save('transactions2.xlsx')